#!/usr/bin/env python3
"""
Convert us_county_data.xlsx to county_data.json format
Expected by Decision 2028 simulator
"""

import openpyxl
import json

def convert_excel_to_json():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook('counties/us_county_data.xlsx')
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    print(f"Found {len(headers)} columns")
    
    # Create county data dictionary with FIPS as key
    county_data = {}
    
    print(f"Processing {ws.max_row - 1} counties...")
    
    for row_idx in range(2, ws.max_row + 1):
        row = list(ws[row_idx])
        row_data = {headers[i]: row[i].value for i in range(len(headers))}
        
        # Get FIPS code (as integer, will be stored as string key)
        fips = str(row_data['FIPS'])
        
        # Calculate vote percentages
        total_votes = row_data['Votes_Avg_Total'] or 0
        if total_votes > 0:
            dem_pct = ((row_data['Votes_Avg_DEM'] or 0) / total_votes) * 100
            rep_pct = ((row_data['Votes_Avg_GOP'] or 0) / total_votes) * 100
            other_pct = ((row_data['Votes_Avg_Other'] or 0) / total_votes) * 100
        else:
            dem_pct = rep_pct = other_pct = 0
        
        # Determine county type based on rural percentage
        rural_pct = row_data['Rural'] or 0
        if rural_pct > 0.7:
            county_type = "Rural"
        elif rural_pct > 0.3:
            county_type = "Mixed"
        else:
            county_type = "Urban"
        
        # Build county object in expected format
        county_obj = {
            "n": row_data['County'],
            "s": row_data['State_Abbrev'],
            "p": int(row_data['Population_Total'] or 0),
            "t": county_type,
            "v": {
                "D": round(dem_pct, 2),
                "R": round(rep_pct, 2),
                "G": 0,  # Green party - will be calculated from Other
                "L": 0,  # Libertarian - will be calculated from Other
                "O": round(other_pct, 2)  # Other parties
            },
            "ig": {
                # Demographics
                "rural": round((row_data['Rural'] or 0) * 100, 2),
                "college": round((row_data['CollegeEducated'] or 0) * 100, 2),
                
                # Racial/Ethnic groups
                "hispanic": round((row_data['Hispanic/Latino'] or 0) * 100, 2),
                "black": round((row_data['AfricanAmerican'] or 0) * 100, 2),
                "asian": round((row_data['Asian'] or 0) * 100, 2),
                "pacific": round((row_data['PacificIslander'] or 0) * 100, 2),
                "native": round((row_data['NativeAmerican'] or 0) * 100, 2),
                
                # Political ideology
                "progressive": round((row_data['Progressive'] or 0) * 100, 2),
                "maga": round((row_data['MAGA'] or 0) * 100, 2),
                "libertarian": round((row_data['Libertarian'] or 0) * 100, 2),
                "centrist": round((row_data['Centrist'] or 0) * 100, 2),
                
                # Religious groups
                "evangelical": round((row_data['Evangelical Protestant'] or 0) * 100, 2),
                "protestant": round((row_data['Other Protestant'] or 0) * 100, 2),
                "catholic": round((row_data['Catholic'] or 0) * 100, 2),
                "christian": round((row_data['Other Christian'] or 0) * 100, 2),
                "jewish": round((row_data['Jewish'] or 0) * 100, 2),
                "muslim": round((row_data['Muslim'] or 0) * 100, 2),
                "other_religion": round((row_data['Other Religion'] or 0) * 100, 2),
                "secular": round((row_data['Religiously Unaffiliated'] or 0) * 100, 2),
                
                # Occupational
                "union": round((row_data['Unionized'] or 0) * 100, 2)
            }
        }
        
        county_data[fips] = county_obj
        
        if row_idx % 500 == 0:
            print(f"  Processed {row_idx - 1} counties...")
    
    print(f"\nTotal counties processed: {len(county_data)}")
    
    # Write to JSON file
    output_path = 'counties/county_data.json'
    print(f"Writing to {output_path}...")
    
    with open(output_path, 'w') as f:
        json.dump(county_data, f, indent=2)
    
    print("✓ Conversion complete!")
    
    # Show sample
    sample_fips = list(county_data.keys())[0]
    print(f"\nSample county (FIPS {sample_fips}):")
    print(json.dumps(county_data[sample_fips], indent=2))

if __name__ == '__main__':
    convert_excel_to_json()
